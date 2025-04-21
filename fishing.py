import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
import os
import json

# 엑셀 파일 읽기
excel_path = "data.xlsx"
fishing_data = "기록"
class_data = "데이터"
wb = load_workbook(excel_path, data_only=True)

# 기록 시트 메인 데이터
sheet_fish = wb[fishing_data]
df_data = []
for row in sheet_fish.iter_rows(min_row=3, min_col=3, max_col=7, values_only=True):
    df_data.append(row)

# 데이터 시트 레벨 데이터
sheet_level = wb[class_data]
ldf_data = []
for row in sheet_level.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
    ldf_data.append(row)

# 기록 시트 거리 보너스 데이터
ddf_data = []
for row in sheet_fish.iter_rows(min_row=6, min_col=10, max_col=11, values_only=True):
    ddf_data.append(row)

# 컬럼 이름 직접 설정
df = [{"수익": row[0], "낚시 시간": row[1], "지역": row[2], "상태": row[3], "도구": row[4]} for row in df_data if row[0] is not None]
ldf = [{"레벨": row[0], "보너스": row[1]} for row in ldf_data if row[0] is not None]
ddf = [{"지역": row[0], "거리": row[1]} for row in ddf_data if row[0] is not None]

# 리스트 추출
region_list = sorted(set(item["지역"] for item in df if item["지역"]))
tool_list = sorted(set(item["도구"] for item in df if item["도구"]))
level_list = list(dict.fromkeys(item["레벨"] for item in ldf if item["레벨"] is not None))
status_list = ["모두", "풍부", "보통", "부족", "고갈"]

# 콤보박스 변경 시 실행될 함수
def update_average(*args):
    selected_region = region_cb.get()
    selected_status = status_cb.get()
    selected_tool = tool_cb.get()
    selected_level = level_cb.get()

    filtered_df = [
        row for row in df
        if (selected_region == "null" or row["지역"] == selected_region)
        and (selected_status == "모두" or selected_status == "null" or row["상태"] == selected_status)
        and (selected_tool == "null" or row["도구"] == selected_tool)
    ]

    # 레벨 보너스 추출
    level_bonus = 0.0
    if selected_level != "null":
        for row in ldf:
            if row["레벨"] == selected_level:
                level_bonus += row["보너스"]
                break

    # 거리 보너스 추출
    distance_bonus = 0.0
    if selected_region != "null":
        for row in ddf:
            if row["지역"] == selected_region:
                distance_bonus += row["거리"]
                break

    # 수익 평균 계산
    if filtered_df:
        total_earnings = sum(row["수익"] for row in filtered_df if row["수익"] is not None)
        total_time = sum(row["낚시 시간"] for row in filtered_df if row["낚시 시간"] is not None)
        if total_time > 0:
            avg = total_earnings / total_time * 100000000 * (1 + 1.3 + level_bonus + distance_bonus)
            result = f"수익 평균: ₩{int(avg):,}"
        else:
            result = "수익 평균: -"
    else:
        result = "수익 평균: -"

    #  평균 수익 값 초기화
    result_label.config(text=result)

# 버튼 클릭시 실행될 함수
def open_excel():
    try:
        os.startfile(excel_path)
    except Exception as e:
        print("엑셀 파일을 찾을 수 없습니다. : ", e)

def on_validate_input(new_value):
    return new_value.isdigit() or new_value == ""

# 레벨 저장 함수
def save_last_setting():
    data = {
        "region": region_cb.get(),
        "status": status_cb.get(),
        "tool": tool_cb.get(),
        "level": level_cb.get()
    }
    with open("fishing_settings.json", "w", encoding="utf-8") as f:
        json.dump(data, f)

# 레벨 불러오기 함수
def load_last_setting():
    try:
        with open("fishing_settings.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            if data.get("region") in region_list:
                region_cb.set(data["region"])
            if data.get("status") in status_list:
                status_cb.set(data["status"])
            if data.get("tool") in tool_list:
                tool_cb.set(data["tool"])
            if data.get("level") in level_list:
                level_cb.set(data["level"])
    except:
        pass    # 파일이 없으면 패스

# 메인 윈도우 생성
icon_path = "fishing.ico"

root = tk.Tk()
root.title("수익 분석 도구")
root.iconbitmap(icon_path)
root.geometry("250x150")  # 창 크기 설정
root.resizable(False, False) # 창 크기 고정
root.protocol("WM_DELETE_WINDOW", lambda: (save_last_setting(), root.destroy()))

# 첫 번째 행: 지역 + 상태
frame1 = tk.Frame(root)
frame1.pack(pady=3)

# 지역
region_label = tk.Label(frame1, text="지역")
region_label.pack(side="left", padx=(0, 5))
region_cb = ttk.Combobox(frame1, values=region_list, width=6)
region_cb.pack(side="left")

if region_list:
    region_cb.set(region_list[0])
else:
    region_cb.set("null")

# 상태
status_label = tk.Label(frame1, text="상태")
status_label.pack(side="left", padx=(15, 5))
status_cb = ttk.Combobox(frame1, values=status_list, width=6)
status_cb.pack(side="left")
status_cb.set(status_list[0])

# 두 번째 행: 도구
frame2 = tk.Frame(root)
frame2.pack(pady=2)

tool_label = tk.Label(frame2, text="도구")
tool_label.pack(side="left", padx=(0, 5))
tool_cb = ttk.Combobox(frame2, values=tool_list, width=14)
tool_cb.pack(side="left")
if tool_list:
    tool_cb.set(tool_list[0])
else:
    tool_cb.set("null")

# 세 번째 행: 레벨과 보너스 합
frame3 = tk.Frame(root)
frame3.pack(pady=2)

level_label = tk.Label(frame3, text="레벨")
level_label.pack(side="left", padx=(0, 5))
level_cb = ttk.Combobox(frame3, values=level_list, width=6)
level_cb.set(level_list[0])
level_cb.pack(side="left")

vcmd = (root.register(on_validate_input), "%P")

# 네 번째 행: 수익 평균
frame4 = tk.Frame(root)
frame4.pack(pady=2)

result_label = tk.Label(frame4, text="수익 평균: ₩0")
result_label.pack()

update_average()

# 다섯 번째 행: 파일 열기
frame5 = tk.Frame(root)
frame5.pack(pady=4)

open_button = tk.Button(frame5, text="엑셀 파일 열기", command=open_excel)
open_button.pack()

# 콤보박스 선택시 평균 가격 업데이트
region_cb.bind("<<ComboboxSelected>>", update_average)
status_cb.bind("<<ComboboxSelected>>", update_average)
tool_cb.bind("<<ComboboxSelected>>", update_average)
level_cb.bind("<<ComboboxSelected>>", update_average)
load_last_setting()

root.mainloop()

